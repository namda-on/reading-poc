#!/usr/bin/env python3
"""리스닝 핑퐁 작성 시트(xlsx) → public/listening-pingpong.data.json 변환.

사용법:
    python3 scripts/xlsx_to_data.py docs/listening-pingpong-template.xlsx

시트 구조는 docs 템플릿의 '설명' 시트 참고. '표현'/'예시'/'대화' 시트를 읽어
프로토타입이 바로 쓰는 JSON을 만든다. 손으로 JSON을 편집하지 말 것 — 시트가 원본.
"""
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "public" / "listening-pingpong.data.json"


def rows(ws):
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        yield {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}


def s(v):
    return "" if v is None else str(v).strip()


def main(path):
    wb = load_workbook(path, data_only=True)

    expressions = {}
    for r in rows(wb["표현"]):
        expressions[s(r["id"])] = {"pattern": s(r["패턴(pattern)"]), "meaning": s(r["뜻(meaning)"])}

    meta = {}
    for r in rows(wb["예시"]):
        eid = s(r["예시ID"])
        v = {"word": s(r["어휘"]), "meaning": s(r["어휘뜻"])}
        if s(r["dictSeq"]):
            v["dictSeq"] = int(float(r["dictSeq"]))
        if s(r["audioUrl"]):
            v["audioUrl"] = s(r["audioUrl"])
        meta[eid] = {
            "vocab": v,
            "expression": s(r["표현id"]),
            "arrange": {"en": s(r["배열_영어"]), "ko": s(r["배열_한글"])},
            "title": s(r["대화제목"]),
        }

    # 대화: 예시ID별로 순서대로 A줄/B문제(연속 B + 같은 순서 = 한 문제)로 묶는다
    turns_by_ex = {}
    for r in rows(wb["대화"]):
        eid = s(r["예시ID"])
        turns_by_ex.setdefault(eid, []).append(r)

    def build_turns(rs):
        turns = []
        i = 0
        while i < len(rs):
            r = rs[i]
            if s(r["화자"]).upper() == "A":
                turns.append({"speaker": "A", "text": s(r["텍스트/선택지"])})
                i += 1
            else:
                order = s(r["순서"])
                opts = []
                while i < len(rs) and s(rs[i]["화자"]).upper() == "B" and s(rs[i]["순서"]) == order:
                    o = rs[i]
                    opt = {"text": s(o["텍스트/선택지"])}
                    if s(o["정답"]).upper() == "O":
                        opt["correct"] = True
                    if s(o["표현"]):
                        opt["expr"] = s(o["표현"])
                    if s(o["어휘강조"]):
                        opt["vocab"] = s(o["어휘강조"])
                    opt["reason"] = s(o["해설"])
                    opts.append(opt)
                    i += 1
                turns.append({"speaker": "B", "options": opts})
        return turns

    examples = []
    for eid, m in meta.items():
        examples.append({
            "vocab": m["vocab"],
            "expression": m["expression"],
            "arrange": m["arrange"],
            "dialogue": {"title": m["title"], "turns": build_turns(turns_by_ex.get(eid, []))},
        })

    data = {"expressions": expressions, "examples": examples}
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {OUT}  ({len(examples)} examples)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: python3 scripts/xlsx_to_data.py <xlsx>")
    main(sys.argv[1])
